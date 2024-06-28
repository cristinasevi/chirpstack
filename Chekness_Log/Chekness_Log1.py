import grpc
from chirpstack_api.as_pb.external import api
from datetime import datetime, timedelta
import time
import openpyxl
import os
import configparser
import re
import json
from subprocess import Popen, PIPE
import threading

server = "localhost:8080"
api_token = "{{token_chirpstack}}"
Com_Count = 0
last_checked_day = None
Device_List = []

# Leer configuraciones desde config.ini
config = configparser.ConfigParser()
config.read('config.ini')

Comm_Threshold = config['DEFAULT'].getint('Comm_Threshold', 300)  # Convertir a entero, usar 300 por defecto si no se encuentra
Log_Rate = config['DEFAULT'].getint('Log_Rate', 60)  # Convertir a entero, usar 60 por defecto si no se encuentra
Plant_Name = config['DEFAULT'].get('Plant_Name', 'Test')  # Si no se encuentra, usar 'Test' por defecto


class Device:
    eui = None
    Last_Seen = None
    Com_Status = None
    Availability = None
    Daily_Com = None
    Weekly_Com = None
    Monthly_Com = None
    RSSI = None 
    instances = []

    def __init__(self, eui, Last_Seen, Com_Status, Availability, Daily_Com, Weekly_Com, Monthly_Com, RSSI):
        self.eui = eui
        self.Last_Seen = Last_Seen
        self.Com_Status = Com_Status
        self.Availability = Availability
        self.Daily_Com = Daily_Com
        self.Weekly_Com = Weekly_Com
        self.Monthly_Com = Monthly_Com
        self.RSSI = RSSI
        Device.instances.append(self)

    def get_eui(self):
        return self.eui
    def get_Last_Seen(self):
        return self.Last_Seen
    def get_Com_Status(self):
        return self.Com_Status
    def get_Availability(self):
        return self.Availability
    def get_Daily_Com(self):
        return self.Daily_Com
    def get_Weekly_Com(self):
        return self.Weekly_Com
    def get_Monthly_Com(self):
        return self.Monthly_Com
    def get_RSSI(self):
        return self.RSSI
    
    def set_eui(self, eui):
        self.eui = eui
    def set_Last_Seen(self, Last_Seen):
        self.Last_Seen = Last_Seen
    def set_Com_Status(self, Com_Status):
        self.Com_Status = Com_Status
    def set_Availability(self, Availability):
        self.Availability = Availability
    def set_Daily_Com(self, Daily_Com):
        self.Daily_Com = Daily_Com
    def set_Weekly_Com(self, Weekly_Com):
        self.Weekly_Com = Weekly_Com
    def set_Monthly_Com(self, Monthly_Com):
        self.Monthly_Com = Monthly_Com
    def set_RSSI(self, RSSI):
        self.RSSI = RSSI

    @classmethod
    def obtener_array_eui(cls):
        return [inst.eui for inst in cls.instances]
    
    @classmethod
    def borrar_dispositivos(cls):
        cls.instances = []
    
def createDevice(eui):
    globals()[eui] = Device(
                        eui=eui,
                        Last_Seen = None,
                        Com_Status = None,
                        Availability = None,
                        Daily_Com = 0,
                        Weekly_Com = 0,
                        Monthly_Com = 0,
                        RSSI = None
                    )

def getChirpstackDevices():
    # The API token (retrieved using the web-interface).
    global api_token
    # Connect without using TLS.
    channel = grpc.insecure_channel(server)

    client = api.ApplicationServiceStub(channel)

    # Define the API key meta-data.
    auth_token = [("authorization", "Bearer %s" % api_token)]

    # Construct request.
    device = api.ListApplicationRequest()
    device.limit = 10000

    resp = client.List(device, metadata=auth_token)

    existent_devices = list()
    for application in resp.result:
        # Device-queue API client.
        client = api.DeviceServiceStub(channel)

        # Define the API key meta-data.
        auth_token = [("authorization", "Bearer %s" % api_token)]

        # Construct request.
        device = api.ListDeviceRequest()
        device.application_id = application.id
        device.limit = 10000

        resp = client.List(device, metadata=auth_token)

        # Print the downlink frame-counter value.
        for device in resp.result:
            existent_devices.append(device.dev_eui)
    return existent_devices

def getActualDevices():
    array_euis = Device.obtener_array_eui()
    return array_euis

def addNewDevices():
    global Device_List
    chirpstack_devices = getChirpstackDevices()
    new_devices = list(set(chirpstack_devices) - set(Device_List))
    for device in new_devices:
        createDevice(device)

def has_day_changed():
    global last_checked_day
    current_day = datetime.now().day
    if last_checked_day is None or current_day != last_checked_day:
        last_checked_day = current_day
        return True
    return False

def getHumanTime(Timestamp):
    HumanTime = datetime.fromtimestamp(Timestamp)
    return HumanTime

def getTimestamp(eui):
    # The API token (retrieved using the web-interface).
    global api_token
    # Connect without using TLS.
    channel = grpc.insecure_channel(server)

    # Device-queue API client.
    client = api.DeviceServiceStub(channel)

    # Define the API key meta-data.
    auth_token = [("authorization", "Bearer %s" % api_token)]

    # Construct request.
    device = api.GetDeviceRequest()
    device.dev_eui = eui
    resp = client.Get(device, metadata=auth_token)
    timestamp = resp.last_seen_at.seconds
    return timestamp

def ComStatus(eui):
    device_timestamp = getTimestamp(eui)
    current_timestamp = time.time()
    time_difference = current_timestamp - device_timestamp
    if time_difference > Comm_Threshold:
        return 0
    else:
        return 1

def updateDevices():
    global Device_List
    Device_List = getActualDevices()
    ValidCount = False
    global Com_Count
    for device in Device_List:
        globals()[device].set_Last_Seen(getHumanTime(getTimestamp(device)))
        globals()[device].set_Com_Status(ComStatus(device))
        if globals()[device].get_Com_Status() == 1:
            globals()[device].set_Daily_Com(globals()[device].get_Daily_Com()+1)
            ValidCount = True
    if ValidCount == True:
        Com_Count = Com_Count + 1
        for device in Device_List:
            Availability = globals()[device].get_Daily_Com()/Com_Count*100
            globals()[device].set_Availability(Availability)

def GetPlantStatus():
    # Connect without using TLS.
    channel = grpc.insecure_channel(server)

    # Device-queue API client.
    client = api.InternalServiceStub(channel)

    # Define the API key meta-data.
    auth_token = [("authorization", "Bearer %s" % api_token)]

    device = api.GetDevicesSummaryRequest()

    resp = client.GetDevicesSummary(device, metadata=auth_token)
    active = resp.active_count
    inactive = resp.inactive_count
    neverseen = resp.never_seen_count
    total = [active,inactive,neverseen]
    return  total

#----------------------------------RSSI---------------------------#
# # Variable global para almacenar valores RSSI
# rssi_values = {}
# rssi_lock = threading.Lock()  # Lock para sincronización de acceso a rssi_values
# total_rssi_count = 0  # Contador global para los valores RSSI capturados

# # Función para extraer el valor RSSI del mensaje
# def filter_and_extract_payload(message):
#     match_rssi = re.search(r'"rssi":\s*(-?\d+)', message)
#     rssi_value = int(match_rssi.group(1)) if match_rssi else None
#     return rssi_value

# # Función para la suscripción MQTT y captura de valores RSSI
# def subscribe_and_capture(device):
#     global total_rssi_count
#     topic = f'application/{{application_id}}/device/{device}/event/up'
    
#     try:
#         process = Popen(['mosquitto_sub', '-h', '{{direccion_ip}}', '-t', topic], stdout=PIPE, stderr=PIPE, universal_newlines=True)
        
#         for line in process.stdout:
#             rssi_value = filter_and_extract_payload(line)
#             if rssi_value is not None:
#                 with rssi_lock:
#                     if total_rssi_count >= 97:
#                         break
#                     rssi_values[device] = rssi_value
#                     total_rssi_count += 1
#                     print(f"Valor RSSI para device {device}: {rssi_value} (Total: {total_rssi_count})")

#                 if total_rssi_count >= 97:
#                     break

#         log_status_to_xlsx(Plant_Name)
#         print(f"Terminada la suscripción para {device}")
        
#     except Exception as e:
#         print(f"Error subscribing to MQTT for {device}: {str(e)}")

#     finally:
#         if 'process' in locals():
#             process.stdout.close()
#             process.stderr.close()
#             process.terminate()
#             process.wait()

# # Función para suscribirse a MQTT en hilos separados para cada dispositivo
# def mqtt_subscribe():
#     global Device_List
#     global rssi_values

#     rssi_values = {device: None for device in Device_List}

#     threads = []
    
#     for device in Device_List:
#         t = threading.Thread(target=subscribe_and_capture, args=(device,))
#         threads.append(t)
#         t.start()

#     for t in threads:
#         t.join()

#     return rssi_values

# # Ejemplo de uso
# rssi_values = mqtt_subscribe()
            
#############################################################

# Variable global para almacenar valores RSSI
rssi_values = {}
rssi_lock = threading.Lock()  # Lock para sincronización de acceso a rssi_values

# Función para extraer el valor RSSI del mensaje
def filter_and_extract_payload(message):
    match_rssi = re.search(r'"rssi":\s*(-?\d+)', message)
    rssi_value = int(match_rssi.group(1)) if match_rssi else None
    return rssi_value

# Función para la suscripción MQTT y captura de valores RSSI
def subscribe_and_capture(device):
    topic = f'application/{{application_id}}/device/{device}/event/up'
    
    try:
        process = Popen(['mosquitto_sub', '-h', '{{direccion_ip}}', '-t', topic], stdout=PIPE, stderr=PIPE, universal_newlines=True)
        
        for line in process.stdout:
            rssi_value = filter_and_extract_payload(line)
            if rssi_value is not None:
                rssi_values[device] = rssi_value
                print(f"Valor RSSI para device {device}: {rssi_value}")

        log_status_to_xlsx(Plant_Name)
        print(f"Terminada la suscripción para {device}")
        
    except Exception as e:
        print(f"Error subscribing to MQTT for {device}: {str(e)}")

    finally:
        if 'process' in locals():
            process.stdout.close()
            process.stderr.close()
            process.terminate()
            process.wait()

# Función para suscribirse a MQTT en hilos separados para cada dispositivo
def mqtt_subscribe():
    global Device_List
    global rssi_values

    rssi_values = {device: None for device in Device_List}

    threads = []
    
    for device in Device_List:
        t = threading.Thread(target=subscribe_and_capture, args=(device,))
        threads.append(t)
        t.start()

    for t in threads:
        t.join()

    return rssi_values

# Ejemplo de uso
rssi_values = mqtt_subscribe()

#---------------------------------------------------------------------------------#

def log_status_to_xlsx(plant_name):
    global Reset
    global Device_List
    global rssi_values

    # Obtener la fecha actual
    current_datetime = datetime.now()
    # Formatear la fecha como "YYYY-MM-DD"
    current_date = current_datetime.strftime("%Y-%m-%d")
    # Crear un nombre de archivo único con el nombre de la planta y la fecha
    filename = f"{plant_name}_{current_date}.xlsx"
    # Abrir el archivo Excel en modo append para agregar datos o crear uno nuevo
    try:
        workbook = openpyxl.load_workbook(filename)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        # Eliminar la hoja de trabajo predeterminada "Sheet"
        default_sheet_name = 'Sheet'
        if default_sheet_name in workbook.sheetnames:
            default_sheet = workbook[default_sheet_name]
            workbook.remove(default_sheet)

    #----------------------------------DATOS COMUNICACIONES---------------------------#
    # Seleccionar la hoja de trabajo Datos_Com o crearla si no existe
    if "Datos_Com" in workbook.sheetnames:
        sheet_com = workbook["Datos_Com"]
    else:
        sheet_com = workbook.create_sheet(title="Datos_Com")

    # Actualizar encabezado
    encabezado = ["Datetime"] + Device_List
    for idx, value in enumerate(encabezado, start=1):
        sheet_com.cell(row=1, column=idx, value=value)

    # Inicializar una fila con el sello de tiempo actual
    row_data_com = [current_datetime]
    for device in Device_List:
        row_data_com.append(globals()[device].get_Com_Status())

    # Agregar la fila a la hoja de trabajo
    sheet_com.append(row_data_com)
    #---------------------------------------------------------------------------------#

    #----------------------------------DATOS DISPONIBILIDAD---------------------------#
    # Seleccionar la hoja de trabajo Datos_Dispo o crearla si no existe
    if "Datos_Dispo" in workbook.sheetnames:
        sheet_avail = workbook["Datos_Dispo"]
    else:
        sheet_avail = workbook.create_sheet(title="Datos_Dispo")
    # Crear un objeto escritor de Excel si el encabezado de Datos_Planta aún no se ha escrito
    if sheet_avail.calculate_dimension() == "A1:A1":
        header_avail = ["EUI", "Dispo %", "Num Com", "15 min"]
        sheet_avail.append(header_avail)

    sheet_to_clear = workbook["Datos_Dispo"]
    sheet_to_clear.delete_rows(2, sheet_to_clear.max_row)

    for device in Device_List:
        row_data_avail = []
        row_data_avail.append(globals()[device].get_eui())
        row_data_avail.append(globals()[device].get_Availability())                      
        row_data_avail.append(globals()[device].get_Daily_Com()) 

        # Calcular si ha comunicado en los últimos 15 minutos
        last_activity_time = globals()[device].get_Last_Seen()  
        has_communicated = (current_datetime - last_activity_time) < timedelta(minutes=15)
        row_data_avail.append(has_communicated)

        sheet_avail.append(row_data_avail)
    #---------------------------------------------------------------------------------#

    #----------------------------------DATOS PLANTA-----------------------------------#
    # Obtener datos de activo, inactivo y nunca visto
    plant_status = GetPlantStatus() 
    # Seleccionar la hoja de trabajo Datos_Planta o crearla si no existe
    if "Datos_Planta" in workbook.sheetnames:
        sheet_plant = workbook["Datos_Planta"]
    else:
        sheet_plant = workbook.create_sheet(title="Datos_Planta")
    # Crear un objeto escritor de Excel si el encabezado de Datos_Planta aún no se ha escrito
    if sheet_plant.calculate_dimension() == "A1:A1":
        header_status = ["Datetime", "Activo", "Inactivo", "Nunca Visto"]
        sheet_plant.append(header_status)
    # Agregar la fila a la hoja de trabajo Status_Log
    sheet_plant.append([current_datetime] + plant_status)
    #---------------------------------------------------------------------------------#

    #----------------------------------DATOS RSSI-----------------------------------#
    #Añadir nueva hoja para RSSI
    if "Datos_RSSI" in workbook.sheetnames:
        sheet_rssi = workbook["Datos_RSSI"]
    else:
        sheet_rssi = workbook.create_sheet(title="Datos_RSSI")

    encabezado_rssi = ["Datetime"] + Device_List
    for idx, value in enumerate(encabezado_rssi, start=1):
        sheet_rssi.cell(row=1, column=idx, value=value)

    #log_values_rssi = [current_datetime] + [rssi_values.get(device, "NaN") for device in Device_List]
    
    log_values_rssi = [current_datetime] + [rssi_values[device] if device in rssi_values and rssi_values[device] is not None else "NaN" for device in Device_List]


    next_row_rssi = sheet_rssi.max_row + 1
    for col_idx, value in enumerate(log_values_rssi, start=1):
        sheet_rssi.cell(row=next_row_rssi, column=col_idx, value=value)

    #---------------------------------------------------------------------------------#

    # Guardar el archivo Excel
    workbook.save(filename)

def guardar_backup():
    global Com_Count
    global Device_List
    global last_checked_day
    # Crear un diccionario con las variables que deseas almacenar
    datos = {"Contador": Com_Count,"Last_Day": last_checked_day ,"Device_List": Device_List}

    # Agregar los valores de Daily_Com de la última iteración al diccionario
    for device in Device_List:
        datos[f"Daily_Com_{device}"] = globals()[device].get_Daily_Com()

    # Agregar los valores de Last_Seen de la última iteración al diccionario
    for device in Device_List:
        datos[f"Last_Seen_{device}"] = globals()[device].get_Last_Seen()

    # Escribir el diccionario en el archivo oculto
    with open(".backup.txt", "w") as f:
        for key, value in datos.items():
            f.write(f"{key}: {value}\n")

def cargar_backup():
    global Com_Count
    global Device_List
    global last_checked_day
    # Intentar cargar datos desde el archivo oculto
    try:
        with open(".backup.txt", "r") as f:
            for line in f:
                key, value = line.strip().split(": ")
                if key == "Contador":
                    Com_Count = int(value)
                elif key == "Last_Day":
                    last_checked_day = int(value)
                elif key == "Device_List":
                    Device_List = eval(value)  # Utiliza eval para convertir la cadena a lista
                elif key.startswith("Daily_Com"):
                    # Extraer el EUI del nombre de la clave y agregar comillas
                    eui = f'{key.replace("Daily_Com_", "")}'
                    createDevice(eui)
                    globals()[eui].set_Daily_Com(int(value))

                elif key.startswith("Last_Seen"):
                    # Extraer el EUI del nombre de la clave y agregar comillas
                    eui = f'{key.replace("Last_Seen_", "")}'
                    last_seen = datetime.strptime(value, '%Y-%m-%d %H:%M:%S')
                    globals()[eui].set_Last_Seen(last_seen)

    except FileNotFoundError:
        Com_Count = 0
        last_checked_day = datetime.now().day
        Device_List = []

def reinciar_datos():
    global Com_Count
    global Device_List
    Com_Count = 0
    Device_List = []

    # Borrar todos los dispositivos almacenados en Device.instances
    Device.borrar_dispositivos()

    try:
        # Eliminar el archivo .backup.txt si existe
        os.remove(".backup.txt")
        print("Archivo de respaldo vaciado correctamente.")
    except FileNotFoundError:
        # No hacer nada si el archivo no existe
        pass

if __name__ == "__main__":
    cargar_backup()
    while True:
        try:
            if has_day_changed() == True:
                reinciar_datos()
                addNewDevices()
                updateDevices()
                guardar_backup()
                log_status_to_xlsx(Plant_Name)
                mqtt_subscribe()


            else:
                addNewDevices()
                updateDevices()
                guardar_backup()
                log_status_to_xlsx(Plant_Name)
                mqtt_subscribe()

        except Exception as e:
            print(f"Error: {e}")
            # Puedes manejar el error según tus necesidades
            # Aquí puedes decidir si quieres salir del bucle o continuar
            # dependiendo del tipo de error.
        # Esperar 30 segundos antes de la siguiente iteración
        time.sleep(Log_Rate)