import re
import json
import subprocess
from concurrent.futures import ThreadPoolExecutor
import openpyxl
from datetime import datetime

def filter_and_extract_rssi(message):
    match_rssi = re.search(r'"rssi":\s*-?\d+', message)
    rssi_part = match_rssi.group() if match_rssi else None
    return rssi_part

def process_eui(eui):
    process = subprocess.Popen(['mosquitto_sub', '-h', 'localhost', '-t', f'application/{{application_id}}/device/{eui}/event/up'], stdout=subprocess.PIPE, stderr=subprocess.PIPE, universal_newlines=True)
    for line in process.stdout:
        rssi_part = filter_and_extract_rssi(line)
        if rssi_part:
            rssi_data = json.loads('{' + rssi_part + '}')
            rssi = rssi_data.get('rssi', None)
            return [datetime.now().strftime('%Y-%m-%d %H:%M:%S'), eui, rssi]

def mqtt_subscribe_and_write_to_excel(euis):
    try:
        wb = openpyxl.load_workbook('device_rssi.xlsx')
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
    if 'rssi' not in wb.sheetnames:
        wb.create_sheet(title='rssi')
    ws = wb['rssi']
    if ws.cell(row=1, column=1).value is None:
        ws.append(['datetime', 'dev_eui', 'rssi'])

    with ThreadPoolExecutor(max_workers=len(euis)) as executor:
        for result in executor.map(process_eui, euis):
            ws.append(result)

    wb.save('device_rssi.xlsx')

if __name__ == "__main__":
    euis = ['{{dev_eui1}}', '{{dev_eui2}}', '{{dev_eui3}}']
    mqtt_subscribe_and_write_to_excel(euis)
