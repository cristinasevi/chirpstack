import re
import json
import subprocess

import openpyxl

def filter_and_extract_payload(message):
    # Utilizamos expresiones regulares para encontrar la parte de la trama que necesitas
    match_corriente = re.search(r'{"corriente1":.+?}', message)
    match_rssi = re.search(r'"rssi":\s*-?\d+', message)
    
    corriente_part = match_corriente.group() if match_corriente else None
    rssi_part = match_rssi.group() if match_rssi else None
     
    return corriente_part, rssi_part

def mqtt_subscribe_and_write_to_excel():
    # Cargar el archivo XLSX existente o crear uno nuevo si no existe
    try:
        wb = openpyxl.load_workbook('device_data.xlsx')
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Eliminar la hoja por defecto si se crea un nuevo archivo

    # Crear una hoja de datos si aún no existe
    if 'device_data' not in wb.sheetnames:
        wb.create_sheet(title='device_data')

    # Seleccionar la hoja de datos
    ws = wb['device_data']

    # Verificar si la segunda fila está vacía y escribir el encabezado
    if ws.cell(row=1, column=1).value is None:
        ws.append(['dev_eui', 'corriente1', 'corriente2', 'corriente3', 'corriente4', 'rssi'])

    
    # Ejecutar mosquitto_sub y leer los mensajes para cada EUI
    process = subprocess.Popen(['mosquitto_sub', '-h', 'localhost', '-t', f'application/1/device/{{dev_eui}}/event/up'], stdout=subprocess.PIPE, stderr=subprocess.PIPE, universal_newlines=True)
    for line in process.stdout:
        # Filtrar y extraer las partes de la trama que necesitas
        corriente_part, rssi_part = filter_and_extract_payload(line)
        
        if corriente_part and rssi_part:
            # Convertir las partes de la trama en diccionarios
            corriente_data = json.loads(corriente_part)
            rssi_data = json.loads('{' + rssi_part + '}')  # Convertir rssi_part en un diccionario
            
            # Obtener los valores de las corrientes y el rssi
            corriente1 = corriente_data.get('corriente1', None)
            corriente2 = corriente_data.get('corriente2', None)
            corriente3 = corriente_data.get('corriente3', None)
            corriente4 = corriente_data.get('corriente4', None)
            rssi = rssi_data.get('rssi', None)
            
            # Obtener la fila siguiente
            row_num = ws.max_row + 1
            
            # Escribir los datos en la hoja de datos
            ws.cell(row=row_num, column=1).value = "0004a30b00fef592"
            ws.cell(row=row_num, column=2).value = corriente1
            ws.cell(row=row_num, column=3).value = corriente2
            ws.cell(row=row_num, column=4).value = corriente3
            ws.cell(row=row_num, column=5).value = corriente4
            ws.cell(row=row_num, column=6).value = rssi
            
            # Guardar los cambios en el archivo XLSX
            wb.save('device_data.xlsx')

if __name__ == "__main__":
    mqtt_subscribe_and_write_to_excel()
