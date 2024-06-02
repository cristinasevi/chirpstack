import re
import json
import subprocess
from concurrent.futures import ThreadPoolExecutor
import openpyxl

def filter_and_extract_payload(message):
    match_corriente = re.search(r'{"corriente1":.+?}', message)
    match_rssi = re.search(r'"rssi":\s*-?\d+', message)
    corriente_part = match_corriente.group() if match_corriente else None
    rssi_part = match_rssi.group() if match_rssi else None
    return corriente_part, rssi_part
 
def process_eui(eui):
    process = subprocess.Popen(['mosquitto_sub', '-h', 'localhost', '-t', f'application/{{application_id}}/device/{eui}/event/up'], stdout=subprocess.PIPE, stderr=subprocess.PIPE, universal_newlines=True)
    for line in process.stdout:
        corriente_part, rssi_part = filter_and_extract_payload(line)
        if corriente_part and rssi_part:
            corriente_data = json.loads(corriente_part)
            rssi_data = json.loads('{' + rssi_part + '}')
            corriente1 = corriente_data.get('corriente1', None)
            corriente2 = corriente_data.get('corriente2', None)
            corriente3 = corriente_data.get('corriente3', None)
            corriente4 = corriente_data.get('corriente4', None)
            rssi = rssi_data.get('rssi', None)
            row = [eui, corriente1, corriente2, corriente3, corriente4, rssi]
            return row

def mqtt_subscribe_and_write_to_excel(euis):
    try:
        wb = openpyxl.load_workbook('device_data.xlsx')
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
    if 'device_data' not in wb.sheetnames:
        wb.create_sheet(title='device_data')
    ws = wb['device_data']
    if ws.cell(row=1, column=1).value is None:
        ws.append(['dev_eui', 'corriente1', 'corriente2', 'corriente3', 'corriente4', 'rssi'])

    with ThreadPoolExecutor(max_workers=len(euis)) as executor:
        for result in executor.map(process_eui, euis):
            ws.append(result)

    wb.save('device_data.xlsx')

if __name__ == "__main__":
    euis = ['0004a30b00fef592', '0004a30b00fef55d', '0004a30b00fef69b']
    mqtt_subscribe_and_write_to_excel(euis)
