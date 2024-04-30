import grpc
from chirpstack_api.as_pb.external import api
from datetime import datetime
import time
import openpyxl
import os
import configparser

server = "localhost:8080"
api_token = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJhcGlfa2V5X2lkIjoiNzYyOGE4NGQtOGRhMy00ZDQ1LWJlNmYtZTM4MWY5MzQ5ZWI1IiwiYXVkIjoiYXMiLCJpc3MiOiJhcyIsIm5iZiI6MTY3NzY4NzI5NCwic3ViIjoiYXBpX2tleSJ9.LBx46H_nzGPkSxqsqVxU_5ig0soMF9dWlsuA6obE1EY"
Com_Count = 0
last_checked_day = None
Device_List = []

# Leer configuraciones desde config.ini
config = configparser.ConfigParser()
config.read('config.ini')

Comm_Threshold = config['DEFAULT'].getint('Comm_Threshold', 300)  # Convertir a entero, usar 300 por defecto si no se encuentra
Log_Rate = config['DEFAULT'].getint('Log_Rate', 60)  # Convertir a entero, usar 60 por defecto si no se encuentra
Plant_Name = config['DEFAULT'].get('Plant_Name', 'Test')  # Si no se encuentra, usar 'Test' por defecto


class Chekness:
    eui = None
    Last_Seen = None
    Com_Status = None
    Availability = None
    Daily_Com = None
    Weekly_Com = None
    Monthly_Com = None
    instances = []

    def __init__(self, eui, Last_Seen, Com_Status, Availability, Daily_Com, Weekly_Com, Monthly_Com):
        self.eui = eui
        self.Last_Seen = Last_Seen
        self.Com_Status = Com_Status
        self.Availability = Availability
        self.Daily_Com = Daily_Com
        self.Weekly_Com = Weekly_Com
        self.Monthly_Com = Monthly_Com
        Chekness.instances.append(self)

    def get_eui(self):
        return self.eui
    def get_Last_Seen(self):
        return self.Last_Seen
    def get_Com_Status(self):
        return self.Com_Status
    def get_Availability(self):
        return self.Availability
    def get_Daily_Com(self):
        return self.Daily_Com
    def get_Weekly_Com(self):
        return self.Weekly_Com
    def get_Monthly_Com(self):
        return self.Monthly_Com
    
    def set_eui(self, eui):
        self.eui = eui
    def set_Last_Seen(self, Last_Seen):
        self.Last_Seen = Last_Seen
    def set_Com_Status(self, Com_Status):
        self.Com_Status = Com_Status
    def set_Availability(self, Availability):
        self.Availability = Availability
    def set_Daily_Com(self, Daily_Com):
        self.Daily_Com = Daily_Com
    def set_Weekly_Com(self, Weekly_Com):
        self.Weekly_Com = Weekly_Com
    def set_Monthly_Com(self, Monthly_Com):
        self.Monthly_Com = Monthly_Com

    @classmethod
    def obtener_array_eui(cls):
        return [inst.eui for inst in cls.instances]
    
    @classmethod
    def borrar_dispositivos(cls):
        cls.instances = []
    
def createDevice(eui):
    globals()[eui] = Chekness(
                        eui=eui,
                        Last_Seen = None,
                        Com_Status = None,
                        Availability = None,
                        Daily_Com = 0,
                        Weekly_Com = 0,
                        Monthly_Com = 0
                    )

def getChirpstackDevices():
    # The API token (retrieved using the web-interface).
    global api_token
    # Connect without using TLS.
    channel = grpc.insecure_channel(server)

    client = api.ApplicationServiceStub(channel)

    # Define the API key meta-data.
    auth_token = [("authorization", "Bearer %s" % api_token)]

    # Construct request.
    device = api.ListApplicationRequest()
    device.limit = 10000

    resp = client.List(device, metadata=auth_token)

    existent_devices = list()
    for application in resp.result:
        # Device-queue API client.
        client = api.DeviceServiceStub(channel)

        # Define the API key meta-data.
        auth_token = [("authorization", "Bearer %s" % api_token)]

        # Construct request.
        device = api.ListDeviceRequest()
        device.application_id = application.id
        device.limit = 10000

        resp = client.List(device, metadata=auth_token)

        # Print the downlink frame-counter value.
        for device in resp.result:
            existent_devices.append(device.dev_eui)
    return existent_devices

def getActualDevices():
    array_euis = Chekness.obtener_array_eui()
    return array_euis

def addNewDevices():
    global Device_List
    chirpstack_devices = getChirpstackDevices()
    new_devices = list(set(chirpstack_devices) - set(Device_List))
    for device in new_devices:
        createDevice(device)

def has_day_changed():
    global last_checked_day
    current_day = datetime.now().day
    if last_checked_day is None or current_day != last_checked_day:
        last_checked_day = current_day
        return True
    return False

def getHumanTime(Timestamp):
    HumanTime = datetime.fromtimestamp(Timestamp)
    return HumanTime

def getTimestamp(eui):
    # The API token (retrieved using the web-interface).
    global api_token
    # Connect without using TLS.
    channel = grpc.insecure_channel(server)

    # Device-queue API client.
    client = api.DeviceServiceStub(channel)

    # Define the API key meta-data.
    auth_token = [("authorization", "Bearer %s" % api_token)]

    # Construct request.
    device = api.GetDeviceRequest()
    device.dev_eui = eui
    resp = client.Get(device, metadata=auth_token)
    timestamp = resp.last_seen_at.seconds
    return timestamp

def ComStatus(eui):
    device_timestamp = getTimestamp(eui)
    current_timestamp = time.time()
    time_difference = current_timestamp - device_timestamp
    if time_difference > Comm_Threshold:
        return 0
    else:
        return 1

def updateDevices():
    global Device_List
    Device_List = getActualDevices()
    ValidCount = False
    global Com_Count
    for device in Device_List:
        globals()[device].set_Last_Seen(getHumanTime(getTimestamp(device)))
        globals()[device].set_Com_Status(ComStatus(device))
        if globals()[device].get_Com_Status() == 1:
            globals()[device].set_Daily_Com(globals()[device].get_Daily_Com()+1)
            ValidCount = True
    if ValidCount == True:
        Com_Count = Com_Count + 1
        for device in Device_List:
            Availability = globals()[device].get_Daily_Com()/Com_Count*100
            globals()[device].set_Availability(Availability)

def GetPlantStatus():
    
    # Connect without using TLS.
    channel = grpc.insecure_channel(server)

    # Device-queue API client.
    client = api.InternalServiceStub(channel)

    # Define the API key meta-data.
    auth_token = [("authorization", "Bearer %s" % api_token)]

    device = api.GetDevicesSummaryRequest()

    resp = client.GetDevicesSummary(device, metadata=auth_token)
    active = resp.active_count
    inactive = resp.inactive_count
    neverseen = resp.never_seen_count
    total = [active,inactive,neverseen]
    return  total

def log_status_to_xlsx(plant_name):
    global Reset
    global Device_List
    # Obtener la fecha actual
    current_datetime = datetime.now()
    # Formatear la fecha como "YYYY-MM-DD"
    current_date = current_datetime.strftime("%Y-%m-%d")
    # Crear un nombre de archivo único con el nombre de la planta y la fecha
    filename = f"{plant_name}_{current_date}.xlsx"
    # Abrir el archivo Excel en modo append para agregar datos o crear uno nuevo
    try:
        workbook = openpyxl.load_workbook(filename)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        # Eliminar la hoja de trabajo predeterminada "Sheet"
        default_sheet_name = 'Sheet'
        if default_sheet_name in workbook.sheetnames:
            default_sheet = workbook[default_sheet_name]
            workbook.remove(default_sheet)

    #----------------------------------DATOS COMUNICACIONES---------------------------#
    # Seleccionar la hoja de trabajo Datos_Com o crearla si no existe
    if "Datos_Com" in workbook.sheetnames:
        sheet_com = workbook["Datos_Com"]
    else:
        sheet_com = workbook.create_sheet(title="Datos_Com")

    # Actualizar encabezado
    encabezado = ["Datetime"] + Device_List
    for idx, value in enumerate(encabezado, start=1):
        sheet_com.cell(row=1, column=idx, value=value)

    # Inicializar una fila con el sello de tiempo actual
    row_data_com = [current_datetime]
    for device in Device_List:
        row_data_com.append(globals()[device].get_Com_Status())

    # Agregar la fila a la hoja de trabajo
    sheet_com.append(row_data_com)
    #---------------------------------------------------------------------------------#

    #----------------------------------DATOS DISPONIBILIDAD---------------------------#
    # Seleccionar la hoja de trabajo Datos_Dispo o crearla si no existe
    if "Datos_Dispo" in workbook.sheetnames:
        sheet_avail = workbook["Datos_Dispo"]
    else:
        sheet_avail = workbook.create_sheet(title="Datos_Dispo")
    # Crear un objeto escritor de Excel si el encabezado de Datos_Planta aún no se ha escrito
    if sheet_avail.calculate_dimension() == "A1:A1":
        header_avail = ["EUI", "Dispo %", "Num Com"]
        sheet_avail.append(header_avail)

    sheet_to_clear = workbook["Datos_Dispo"]
    sheet_to_clear.delete_rows(2, sheet_to_clear.max_row)

    for device in Device_List:
        row_data_avail = []
        row_data_avail.append(globals()[device].get_eui())
        row_data_avail.append(globals()[device].get_Availability())                      
        row_data_avail.append(globals()[device].get_Daily_Com())                      
        sheet_avail.append(row_data_avail)
    
    #---------------------------------------------------------------------------------#

    #----------------------------------DATOS PLANTA-----------------------------------#
    # Obtener datos de activo, inactivo y nunca visto
    plant_status = GetPlantStatus() 
    # Seleccionar la hoja de trabajo Datos_Planta o crearla si no existe
    if "Datos_Planta" in workbook.sheetnames:
        sheet_plant = workbook["Datos_Planta"]
    else:
        sheet_plant = workbook.create_sheet(title="Datos_Planta")
    # Crear un objeto escritor de Excel si el encabezado de Datos_Planta aún no se ha escrito
    if sheet_plant.calculate_dimension() == "A1:A1":
        header_status = ["Datetime", "Activo", "Inactivo", "Nunca Visto"]
        sheet_plant.append(header_status)
    # Agregar la fila a la hoja de trabajo Status_Log
    sheet_plant.append([current_datetime] + plant_status)
    #---------------------------------------------------------------------------------#

    # Guardar el archivo Excel
    workbook.save(filename)

def guardar_backup():
    global Com_Count
    global Device_List
    global last_checked_day
    # Crear un diccionario con las variables que deseas almacenar
    datos = {"Contador": Com_Count,"Last_Day": last_checked_day ,"Device_List": Device_List}

    # Agregar los valores de Daily_Com de la última iteración al diccionario
    for device in Device_List:
        datos[f"Daily_Com_{device}"] = globals()[device].get_Daily_Com()

    # Agregar los valores de Last_Seen de la última iteración al diccionario
    for device in Device_List:
        datos[f"Last_Seen_{device}"] = globals()[device].get_Last_Seen()

    # Escribir el diccionario en el archivo oculto
    with open(".backup.txt", "w") as f:
        for key, value in datos.items():
            f.write(f"{key}: {value}\n")

def cargar_backup():
    global Com_Count
    global Device_List
    global last_checked_day
    # Intentar cargar datos desde el archivo oculto
    try:
        with open(".backup.txt", "r") as f:
            for line in f:
                key, value = line.strip().split(": ")
                if key == "Contador":
                    Com_Count = int(value)
                elif key == "Last_Day":
                    last_checked_day = int(value)
                elif key == "Device_List":
                    Device_List = eval(value)  # Utiliza eval para convertir la cadena a lista
                elif key.startswith("Daily_Com"):
                    # Extraer el EUI del nombre de la clave y agregar comillas
                    eui = f'{key.replace("Daily_Com_", "")}'
                    createDevice(eui)
                    globals()[eui].set_Daily_Com(int(value))

                elif key.startswith("Last_Seen"):
                    # Extraer el EUI del nombre de la clave y agregar comillas
                    eui = f'{key.replace("Last_Seen_", "")}'
                    last_seen = datetime.strptime(value, '%Y-%m-%d %H:%M:%S')
                    globals()[eui].set_Last_Seen(last_seen)

    except FileNotFoundError:
        Com_Count = 0
        last_checked_day = datetime.now().day
        Device_List = []

def reinciar_datos():
    global Com_Count
    global Device_List
    Com_Count = 0
    Device_List = []

    # Borrar todos los dispositivos almacenados en Chekness.instances
    Chekness.borrar_dispositivos()

    try:
        # Eliminar el archivo .backup.txt si existe
        os.remove(".backup.txt")
        print("Archivo de respaldo vaciado correctamente.")
    except FileNotFoundError:
        # No hacer nada si el archivo no existe
        pass

if __name__ == "__main__":
    cargar_backup()
    while True:
        try:
            if has_day_changed() == True:
                reinciar_datos()
                addNewDevices()
                updateDevices()
                guardar_backup()
                log_status_to_xlsx(Plant_Name)
                

            else:
                addNewDevices()
                updateDevices()
                guardar_backup()
                log_status_to_xlsx(Plant_Name)

        except Exception as e:
            print(f"Error: {e}")
            # Puedes manejar el error según tus necesidades
            # Aquí puedes decidir si quieres salir del bucle o continuar
            # dependiendo del tipo de error.
        # Esperar 30 segundos antes de la siguiente iteración
        time.sleep(Log_Rate)



